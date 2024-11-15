import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill


df = pd.read_excel('C:/Users/Excalibur/PycharmProjects/Rosatom1/prompt_experiments/output.xlsx')
df1 = pd.read_excel('C:/Users/Excalibur/PycharmProjects/Rosatom1/data/размеченный.xlsx')
df["Product"] = df["Product"].astype(str).str.lower().str.strip()
df1["Product"] = df1["Product"].astype(str).str.lower().str.strip()
df = df['Product']
df1 = df1['Product']
a = sorted(set(df) & set(df1))
print(len(a)/len(df) * 100)
print(len(a)/len(df1) * 100)
wb = load_workbook(filename='C:/Users/Excalibur/PycharmProjects/Rosatom1/prompt_experiments/output.xlsx')
ws = wb.active
white_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
for row in ws.iter_rows():
    for cell in row:
        cell.fill = white_fill

wb.save('C:/Users/Excalibur/PycharmProjects/Rosatom1/prompt_experiments/output.xlsx')

wb = load_workbook(filename='C:/Users/Excalibur/PycharmProjects/Rosatom1/data/размеченный.xlsx')
ws = wb.active
for row in ws.iter_rows():
    for cell in row:
        cell.fill = white_fill
wb.save('C:/Users/Excalibur/PycharmProjects/Rosatom1/data/размеченный.xlsx')

wb = load_workbook(filename='C:/Users/Excalibur/PycharmProjects/Rosatom1/data/размеченный.xlsx')
ws = wb.active
green = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")

wb = load_workbook(filename='C:/Users/Excalibur/PycharmProjects/Rosatom1/data/размеченный.xlsx')
sheet = wb.active
for row in sheet.iter_rows(min_row=2, min_col=2, max_col=2, max_row=len(df)):
    for cell in row:
        try:
            if cell.value.lower().strip() in a:
                cell.fill = green
        except:
            pass

wb.save('C:/Users/Excalibur/PycharmProjects/Rosatom1/data/размеченный.xlsx')

wb = load_workbook(filename='C:/Users/Excalibur/PycharmProjects/Rosatom1/prompt_experiments/output.xlsx')
sheet = wb.active
for row in sheet.iter_rows(min_row=2, min_col=2, max_col=2, max_row=len(df)):
    for cell in row:
        try:
            if cell.value.lower().strip() in a:
                cell.fill = green
        except:
            pass
wb.save('C:/Users/Excalibur/PycharmProjects/Rosatom1/prompt_experiments/output.xlsx')










